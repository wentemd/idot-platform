"""
IDOT Bid Intelligence Platform - API Routes
Routes for the flat bids table schema with security hardening
"""
from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse
from typing import Optional, List
from slowapi import Limiter
from slowapi.util import get_remote_address
import sqlite3
import os
import io

from app.api.users import (
    get_user_by_token, check_and_reset_daily_searches, 
    increment_daily_searches, get_user_limits
)

router = APIRouter()

# Rate limiter - will use the one from main.py app state
limiter = Limiter(key_func=get_remote_address)


def get_current_user(request: Request) -> Optional[dict]:
    """Get current user from session token"""
    token = request.cookies.get("session_token")
    if not token:
        return None
    return get_user_by_token(token)


def check_search_limit(request: Request) -> dict:
    """Check if user can perform a search, return user limits"""
    user = get_current_user(request)
    
    if not user:
        # Anonymous user - very limited
        return {
            'user': None,
            'can_search': True,  # Allow some anonymous searches
            'results_limit': 25,
            'is_pro': False
        }
    
    limits = get_user_limits(user)
    
    # Pro users - unlimited
    if limits['daily_searches'] > 1000:
        return {
            'user': user,
            'can_search': True,
            'results_limit': limits['results_per_query'],
            'is_pro': True
        }
    
    # Free users - check daily limit
    current_searches = check_and_reset_daily_searches(user['id'])
    
    if current_searches >= limits['daily_searches']:
        return {
            'user': user,
            'can_search': False,
            'results_limit': limits['results_per_query'],
            'is_pro': False,
            'searches_remaining': 0
        }
    
    return {
        'user': user,
        'can_search': True,
        'results_limit': limits['results_per_query'],
        'is_pro': False,
        'searches_remaining': limits['daily_searches'] - current_searches
    }

router = APIRouter()

def get_db():
    """Get database connection"""
    db_path = os.getenv("DATABASE_PATH", "/app/data/idot_intelligence.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

# ============================================================================
# STATS / ANALYTICS ENDPOINTS
# ============================================================================

@router.get("/stats")
@limiter.limit("60/minute")
async def get_stats(request: Request):
    """Get overall database statistics"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) FROM bids")
    total_bids = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT contract_number) FROM bids")
    total_contracts = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT bidder_name) FROM bids")
    total_contractors = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT item_number) FROM bids")
    total_items = cursor.fetchone()[0]
    
    conn.close()
    
    return {
        "total_bids": total_bids,
        "total_contracts": total_contracts,
        "total_contractors": total_contractors,
        "total_items": total_items
    }


@router.get("/analytics/summary")
@limiter.limit("60/minute")
async def get_analytics_summary(request: Request):
    """Get comprehensive analytics summary for dashboard"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Total bid rows
    cursor.execute("SELECT COUNT(*) FROM bids")
    total_bid_rows = cursor.fetchone()[0]
    
    # Unique contracts
    cursor.execute("SELECT COUNT(DISTINCT contract_number) FROM bids")
    unique_contracts = cursor.fetchone()[0]
    
    # Unique contractors
    cursor.execute("SELECT COUNT(DISTINCT bidder_name) FROM bids")
    unique_contractors = cursor.fetchone()[0]
    
    # Unique items
    cursor.execute("SELECT COUNT(DISTINCT item_number) FROM bids")
    unique_items = cursor.fetchone()[0]
    
    # Unique counties
    cursor.execute("SELECT COUNT(DISTINCT county) FROM bids WHERE county IS NOT NULL AND county != ''")
    unique_counties = cursor.fetchone()[0]
    
    # Year range
    cursor.execute("""
        SELECT 
            MIN(CAST(substr(letting_date, length(letting_date)-3) AS INTEGER)) as min_year,
            MAX(CAST(substr(letting_date, length(letting_date)-3) AS INTEGER)) as max_year
        FROM bids
        WHERE letting_date IS NOT NULL
    """)
    year_row = cursor.fetchone()
    
    conn.close()
    
    return {
        "total_bid_rows": total_bid_rows,
        "unique_contracts": unique_contracts,
        "unique_contractors": unique_contractors,
        "unique_items": unique_items,
        "unique_counties": unique_counties,
        "year_range": {
            "min": year_row['min_year'] if year_row else None,
            "max": year_row['max_year'] if year_row else None
        }
    }


@router.get("/health")
@limiter.limit("60/minute")
async def health_check(request: Request):
    """Health check endpoint"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        conn.close()
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "error": str(e)}

# ============================================================================
# PAY ITEM SEARCH
# ============================================================================

@router.get("/search/pay-item/{item_number}")
@limiter.limit("30/minute")
async def search_pay_item(
    request: Request,
    item_number: str,
    county: Optional[str] = None,
    district: Optional[str] = None,
    year_start: Optional[int] = None,
    year_end: Optional[int] = None,
    limit: int = Query(default=500, le=1000)
):
    """
    Search for a pay item and get pricing history from ALL bidders.
    Returns unit prices, quantities, and yearly trends with WEIGHTED averages.
    """
    # Check search limits
    search_check = check_search_limit(request)
    if not search_check['can_search']:
        raise HTTPException(
            status_code=429, 
            detail="Daily search limit reached. Upgrade to Pro for unlimited searches."
        )
    
    # Apply results limit based on tier
    effective_limit = min(limit, search_check['results_limit'])
    
    # Increment search count for logged-in free users
    if search_check['user'] and not search_check['is_pro']:
        increment_daily_searches(search_check['user']['id'])
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Build query
    query = """
        SELECT 
            item_number,
            item_description,
            contract_number,
            letting_date,
            substr(letting_date, length(letting_date)-3) as letting_year,
            county,
            district,
            bidder_name,
            bidder_rank,
            is_winner,
            quantity,
            unit,
            unit_price,
            extension,
            engineers_est_unit_price,
            is_low_item,
            item_rank
        FROM bids
        WHERE item_number LIKE ?
    """
    params = [f"%{item_number}%"]
    
    if county:
        query += " AND county LIKE ?"
        params.append(f"%{county}%")
    
    if district:
        query += " AND district LIKE ?"
        params.append(f"%{district}%")
    
    if year_start:
        query += " AND CAST(substr(letting_date, length(letting_date)-3) AS INTEGER) >= ?"
        params.append(year_start)
    
    if year_end:
        query += " AND CAST(substr(letting_date, length(letting_date)-3) AS INTEGER) <= ?"
        params.append(year_end)
    
    query += " ORDER BY letting_date DESC, contract_number, bidder_rank LIMIT ?"
    params.append(effective_limit)
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    
    # Get yearly statistics with WEIGHTED averages - WINNING BIDS ONLY
    stats_query = """
        SELECT 
            substr(letting_date, length(letting_date)-3) as year,
            COUNT(*) as bid_count,
            SUM(extension) / NULLIF(SUM(quantity), 0) as weighted_avg_price,
            MIN(unit_price) as min_price,
            MAX(unit_price) as max_price,
            SUM(quantity) as total_quantity,
            SUM(extension) as total_value
        FROM bids
        WHERE item_number LIKE ?
        AND unit_price > 0
        AND quantity > 0
        AND is_winner = 'Y'
    """
    stats_params = [f"%{item_number}%"]
    
    if county:
        stats_query += " AND county LIKE ?"
        stats_params.append(f"%{county}%")
    
    if district:
        stats_query += " AND district LIKE ?"
        stats_params.append(f"%{district}%")
    
    stats_query += " GROUP BY substr(letting_date, length(letting_date)-3) ORDER BY year"
    
    cursor.execute(stats_query, stats_params)
    yearly_stats = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return {
        "item_number": item_number,
        "filters": {"county": county, "district": district, "year_start": year_start, "year_end": year_end},
        "result_count": len(rows),
        "yearly_trends": yearly_stats,
        "bids": [dict(row) for row in rows]
    }


@router.get("/search/pay-item-exact/{item_number}")
@limiter.limit("30/minute")
async def search_pay_item_exact(
    request: Request,
    item_number: str,
    limit: int = Query(default=200, le=500)
):
    """Exact match search for a specific pay item code"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            item_number,
            item_description,
            contract_number,
            letting_date,
            county,
            district,
            bidder_name,
            bidder_rank,
            is_winner,
            quantity,
            unit,
            unit_price,
            extension
        FROM bids
        WHERE item_number = ?
        ORDER BY letting_date DESC
        LIMIT ?
    """, [item_number, limit])
    
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "item_number": item_number,
        "result_count": len(rows),
        "bids": [dict(row) for row in rows]
    }

# ============================================================================
# CONTRACTOR SEARCH
# ============================================================================

@router.get("/search/contractor/{name}")
@limiter.limit("30/minute")
async def search_contractor(
    request: Request,
    name: str,
    county: Optional[str] = None,
    district: Optional[str] = None,
    year_start: Optional[int] = None,
    year_end: Optional[int] = None
):
    """
    Search contractor bidding history.
    Returns contract-level summary with win rates.
    """
    # Check search limits
    search_check = check_search_limit(request)
    if not search_check['can_search']:
        raise HTTPException(
            status_code=429, 
            detail="Daily search limit reached. Upgrade to Pro for unlimited searches."
        )
    
    # Increment search count for logged-in free users
    if search_check['user'] and not search_check['is_pro']:
        increment_daily_searches(search_check['user']['id'])
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Build WHERE clause for filters
    where_clause = "WHERE bidder_name LIKE ?"
    params = [f"%{name}%"]
    
    if county:
        where_clause += " AND county LIKE ?"
        params.append(f"%{county}%")
    
    if district:
        where_clause += " AND district LIKE ?"
        params.append(f"%{district}%")
    
    if year_start:
        where_clause += " AND CAST(substr(letting_date, length(letting_date)-3) AS INTEGER) >= ?"
        params.append(year_start)
    
    if year_end:
        where_clause += " AND CAST(substr(letting_date, length(letting_date)-3) AS INTEGER) <= ?"
        params.append(year_end)
    
    # Get contract-level summary (one row per contract)
    contracts_query = f"""
        SELECT 
            contract_number,
            letting_date,
            county,
            district,
            bidder_name,
            bidder_rank,
            total_bid_amount,
            bid_spread_pct,
            is_winner,
            COUNT(*) as item_count
        FROM bids
        {where_clause}
        GROUP BY contract_number, bidder_name
        ORDER BY letting_date DESC
    """
    
    cursor.execute(contracts_query, params)
    contracts = [dict(row) for row in cursor.fetchall()]
    
    # Get win statistics
    stats_query = f"""
        SELECT 
            bidder_name,
            COUNT(DISTINCT contract_number) as contracts_bid,
            COUNT(DISTINCT CASE WHEN is_winner = 'Y' THEN contract_number END) as contracts_won,
            ROUND(100.0 * COUNT(DISTINCT CASE WHEN is_winner = 'Y' THEN contract_number END) / 
                COUNT(DISTINCT contract_number), 1) as win_rate,
            ROUND(AVG(bidder_rank), 1) as avg_rank
        FROM bids
        {where_clause}
        GROUP BY bidder_name
    """
    
    cursor.execute(stats_query, params)
    stats_rows = cursor.fetchall()
    
    # Get total won value separately
    won_query = f"""
        SELECT SUM(total_bid_amount) as total_won_value
        FROM (
            SELECT DISTINCT contract_number, total_bid_amount
            FROM bids
            {where_clause} AND is_winner = 'Y'
        )
    """
    
    cursor.execute(won_query, params)
    won_value_row = cursor.fetchone()
    total_won_value = won_value_row['total_won_value'] if won_value_row and won_value_row['total_won_value'] else 0
    
    stats = []
    for row in stats_rows:
        stat = dict(row)
        stat['total_won_value'] = total_won_value
        stats.append(stat)
    
    conn.close()
    
    return {
        "search_term": name,
        "filters": {"county": county, "district": district, "year_start": year_start, "year_end": year_end},
        "contractor_stats": stats,
        "contract_count": len(contracts),
        "contracts": contracts
    }

# ============================================================================
# CONTRACT SEARCH
# ============================================================================

@router.get("/search/contract/{contract_number}")
@limiter.limit("30/minute")
async def search_contract(request: Request, contract_number: str):
    """
    Get all bids for a specific contract.
    Returns data organized for item-by-item comparison across bidders.
    """
    # Check search limits
    search_check = check_search_limit(request)
    if not search_check['can_search']:
        raise HTTPException(
            status_code=429, 
            detail="Daily search limit reached. Upgrade to Pro for unlimited searches."
        )
    
    # Increment search count for logged-in free users
    if search_check['user'] and not search_check['is_pro']:
        increment_daily_searches(search_check['user']['id'])
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT *
        FROM bids
        WHERE contract_number LIKE ?
        ORDER BY item_number, bidder_rank
    """, [f"%{contract_number}%"])
    
    rows = cursor.fetchall()
    conn.close()
    
    if not rows:
        raise HTTPException(status_code=404, detail="Contract not found")
    
    bids = [dict(row) for row in rows]
    
    # Organize by item for comparison view
    items_comparison = {}
    bidders_info = {}
    
    for bid in bids:
        item_num = bid['item_number']
        bidder = bid['bidder_name']
        
        # Track bidder info
        if bidder not in bidders_info:
            bidders_info[bidder] = {
                'rank': bid['bidder_rank'],
                'total_bid': bid['total_bid_amount'],
                'is_winner': bid['is_winner']
            }
        
        # Organize items
        if item_num not in items_comparison:
            items_comparison[item_num] = {
                'item_number': item_num,
                'item_description': bid['item_description'],
                'quantity': bid['quantity'],
                'unit': bid['unit'],
                'engineers_estimate': bid.get('engineers_est_unit_price'),
                'bidder_prices': {}
            }
        
        items_comparison[item_num]['bidder_prices'][bidder] = {
            'unit_price': bid['unit_price'],
            'extension': bid['extension'],
            'is_winner': bid['is_winner']
        }
    
    # Sort bidders by rank
    sorted_bidders = sorted(bidders_info.items(), key=lambda x: x[1]['rank'])
    
    return {
        "contract_number": contract_number,
        "result_count": len(bids),
        "bidders": [{"name": name, **info} for name, info in sorted_bidders],
        "items_comparison": list(items_comparison.values()),
        "bids": bids  # Keep raw bids for backward compatibility
    }

# ============================================================================
# PRICING ANALYTICS
# ============================================================================

@router.get("/pricing/item-summary")
@limiter.limit("20/minute")
async def get_item_pricing_summary(
    request: Request,
    min_occurrences: int = Query(default=10, description="Minimum bid count to include"),
    limit: int = Query(default=50, le=100)
):
    """
    Get pricing summary for all items with sufficient data.
    Includes WEIGHTED average prices from WINNING BIDS, price ranges, and bid counts.
    """
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            item_number,
            item_description,
            unit,
            COUNT(*) as bid_count,
            COUNT(DISTINCT contract_number) as contract_count,
            SUM(extension) / NULLIF(SUM(quantity), 0) as weighted_avg_price,
            MIN(unit_price) as min_price,
            MAX(unit_price) as max_price,
            SUM(quantity) as total_quantity
        FROM bids
        WHERE unit_price > 0 AND quantity > 0
        AND is_winner = 'Y'
        GROUP BY item_number, item_description, unit
        HAVING COUNT(*) >= ?
        ORDER BY bid_count DESC
        LIMIT ?
    """, [min_occurrences, limit])
    
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "min_occurrences": min_occurrences,
        "result_count": len(rows),
        "items": [dict(row) for row in rows]
    }


@router.get("/pricing/county-comparison/{item_number}")
@limiter.limit("20/minute")
async def get_county_comparison(request: Request, item_number: str):
    """
    Compare pricing for an item across counties with WEIGHTED averages from WINNING BIDS.
    """
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            county,
            COUNT(*) as bid_count,
            SUM(extension) / NULLIF(SUM(quantity), 0) as weighted_avg_price,
            MIN(unit_price) as min_price,
            MAX(unit_price) as max_price
        FROM bids
        WHERE item_number LIKE ?
        AND unit_price > 0
        AND quantity > 0
        AND is_winner = 'Y'
        GROUP BY county
        HAVING COUNT(*) >= 3
        ORDER BY weighted_avg_price DESC
    """, [f"%{item_number}%"])
    
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "item_number": item_number,
        "counties": [dict(row) for row in rows]
    }


@router.get("/pricing/district-comparison/{item_number}")
@limiter.limit("20/minute")
async def get_district_comparison(request: Request, item_number: str):
    """
    Compare pricing for an item across districts with WEIGHTED averages from WINNING BIDS.
    """
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            district,
            COUNT(*) as bid_count,
            SUM(extension) / NULLIF(SUM(quantity), 0) as weighted_avg_price,
            MIN(unit_price) as min_price,
            MAX(unit_price) as max_price,
            SUM(quantity) as total_quantity
        FROM bids
        WHERE item_number LIKE ?
        AND unit_price > 0
        AND quantity > 0
        AND district IS NOT NULL
        AND is_winner = 'Y'
        GROUP BY district
        HAVING COUNT(*) >= 3
        ORDER BY weighted_avg_price DESC
    """, [f"%{item_number}%"])
    
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "item_number": item_number,
        "districts": [dict(row) for row in rows]
    }

# ============================================================================
# BROWSE / LIST ENDPOINTS
# ============================================================================

@router.get("/browse/items")
@limiter.limit("30/minute")
async def browse_items(
    request: Request,
    search: Optional[str] = None,
    limit: int = Query(default=50, le=100)
):
    """Browse all pay items with optional search - uses WEIGHTED averages from WINNING BIDS"""
    conn = get_db()
    cursor = conn.cursor()
    
    if search:
        cursor.execute("""
            SELECT 
                item_number,
                item_description,
                unit,
                COUNT(*) as bid_count,
                COUNT(DISTINCT contract_number) as contract_count,
                ROUND(SUM(extension) / NULLIF(SUM(quantity), 0), 2) as avg_price
            FROM bids
            WHERE (item_number LIKE ? OR item_description LIKE ?)
            AND unit_price > 0
            AND quantity > 0
            AND is_winner = 'Y'
            GROUP BY item_number, item_description, unit
            ORDER BY bid_count DESC
            LIMIT ?
        """, [f"%{search}%", f"%{search}%", limit])
    else:
        cursor.execute("""
            SELECT 
                item_number,
                item_description,
                unit,
                COUNT(*) as bid_count,
                COUNT(DISTINCT contract_number) as contract_count,
                ROUND(SUM(extension) / NULLIF(SUM(quantity), 0), 2) as avg_price
            FROM bids
            WHERE unit_price > 0
            AND quantity > 0
            AND is_winner = 'Y'
            GROUP BY item_number, item_description, unit
            ORDER BY bid_count DESC
            LIMIT ?
        """, [limit])
    
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "search": search,
        "result_count": len(rows),
        "items": [dict(row) for row in rows]
    }


@router.get("/browse/contractors")
@limiter.limit("30/minute")
async def browse_contractors(
    request: Request,
    search: Optional[str] = None,
    limit: int = Query(default=50, le=100)
):
    """Browse all contractors with optional search"""
    conn = get_db()
    cursor = conn.cursor()
    
    base_query = """
        SELECT 
            bidder_name,
            COUNT(DISTINCT contract_number) as contracts_bid,
            COUNT(DISTINCT CASE WHEN is_winner = 'Y' THEN contract_number END) as contracts_won,
            ROUND(100.0 * COUNT(DISTINCT CASE WHEN is_winner = 'Y' THEN contract_number END) / 
                COUNT(DISTINCT contract_number), 1) as win_rate
        FROM bids
    """
    
    if search:
        base_query += " WHERE bidder_name LIKE ?"
        base_query += " GROUP BY bidder_name ORDER BY contracts_bid DESC LIMIT ?"
        cursor.execute(base_query, [f"%{search}%", limit])
    else:
        base_query += " GROUP BY bidder_name ORDER BY contracts_bid DESC LIMIT ?"
        cursor.execute(base_query, [limit])
    
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "search": search,
        "result_count": len(rows),
        "contractors": [dict(row) for row in rows]
    }


@router.get("/browse/contracts")
@limiter.limit("30/minute")
async def browse_contracts(
    request: Request,
    county: Optional[str] = None,
    district: Optional[str] = None,
    year: Optional[int] = None,
    limit: int = Query(default=50, le=100)
):
    """Browse contracts with optional filters"""
    conn = get_db()
    cursor = conn.cursor()
    
    query = """
        SELECT DISTINCT
            contract_number,
            letting_date,
            county,
            district,
            num_bidders,
            engineers_estimate,
            awarded
        FROM bids
        WHERE 1=1
    """
    params = []
    
    if county:
        query += " AND county LIKE ?"
        params.append(f"%{county}%")
    
    if district:
        query += " AND district LIKE ?"
        params.append(f"%{district}%")
    
    if year:
        query += " AND CAST(substr(letting_date, length(letting_date)-3) AS INTEGER) = ?"
        params.append(year)
    
    query += " ORDER BY letting_date DESC LIMIT ?"
    params.append(limit)
    
    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "filters": {"county": county, "district": district, "year": year},
        "result_count": len(rows),
        "contracts": [dict(row) for row in rows]
    }


@router.get("/browse/districts")
@limiter.limit("60/minute")
async def browse_districts(request: Request):
    """Get list of all districts"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT DISTINCT district, COUNT(DISTINCT contract_number) as contract_count
        FROM bids
        WHERE district IS NOT NULL AND district != ''
        GROUP BY district
        ORDER BY district
    """)
    
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "districts": [dict(row) for row in rows]
    }


@router.get("/browse/counties")
@limiter.limit("60/minute")
async def browse_counties(request: Request):
    """Get list of all counties"""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT DISTINCT county, COUNT(DISTINCT contract_number) as contract_count
        FROM bids
        WHERE county IS NOT NULL AND county != ''
        GROUP BY county
        ORDER BY county
    """)
    
    rows = cursor.fetchall()
    conn.close()
    
    return {
        "counties": [dict(row) for row in rows]
    }


# ============================================================================
# ESTIMATOR EXCEL UPLOAD
# ============================================================================

@router.post("/estimator/price-items")
@limiter.limit("10/minute")
async def price_items_from_excel(
    request: Request,
    file: UploadFile = File(...),
    districts: str = Form(default=""),
    year_start: Optional[int] = Form(default=None),
    year_end: Optional[int] = Form(default=None)
):
    """
    Upload an Excel file with item numbers and quantities.
    Returns the file with weighted average prices filled in.
    
    PRO FEATURE - requires active subscription.
    
    Expected Excel format:
    - Column A: Item Number (required)
    - Column B: Item Description (will be filled if empty)
    - Column C: Quantity (required for extension calc)
    - Column D: Unit (will be filled if empty)
    - Column E: Unit Price (will be filled by system)
    - Column F: Extension (will be calculated)
    
    Limited to 300 items to prevent bulk data extraction.
    """
    # Check if user has Pro access
    user = get_current_user(request)
    if not user:
        raise HTTPException(
            status_code=401, 
            detail="Please log in to use the Estimator Tool"
        )
    
    limits = get_user_limits(user)
    if not limits['estimator_access']:
        raise HTTPException(
            status_code=403, 
            detail="The Estimator Tool requires a Pro subscription. Start your 7-day free trial today!"
        )
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Read the uploaded file
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")
    
    # Parse districts
    district_list = [d.strip() for d in districts.split(',') if d.strip()] if districts else []
    
    # Find header row and data start
    header_row = 1
    data_start = 2
    
    # Check if first row looks like headers
    first_cell = str(ws.cell(row=1, column=1).value or "").lower()
    if 'item' in first_cell or 'number' in first_cell or 'code' in first_cell:
        header_row = 1
        data_start = 2
    else:
        # No header, data starts at row 1
        header_row = 0
        data_start = 1
    
    # Collect item numbers from column A
    items_to_price = []
    row = data_start
    while row <= ws.max_row and len(items_to_price) < 300:
        item_num = ws.cell(row=row, column=1).value
        if item_num:
            item_num = str(item_num).strip()
            if item_num:
                quantity = ws.cell(row=row, column=3).value
                try:
                    quantity = float(quantity) if quantity else 0
                except:
                    quantity = 0
                items_to_price.append({
                    'row': row,
                    'item_number': item_num,
                    'quantity': quantity
                })
        row += 1
    
    if len(items_to_price) == 0:
        raise HTTPException(status_code=400, detail="No item numbers found in column A")
    
    if row <= ws.max_row:
        # More than 300 items
        raise HTTPException(
            status_code=400, 
            detail=f"File contains more than 300 items. Please limit to 300 items per upload. Found items up to row {row}."
        )
    
    # Get pricing from database
    conn = get_db()
    cursor = conn.cursor()
    
    # Build WHERE clause for districts and years
    district_clause = ""
    year_clause = ""
    
    if district_list:
        placeholders = ','.join(['?' for _ in district_list])
        district_clause = f" AND district IN ({placeholders})"
    
    if year_start:
        year_clause += f" AND CAST(substr(letting_date, length(letting_date)-3) AS INTEGER) >= {year_start}"
    if year_end:
        year_clause += f" AND CAST(substr(letting_date, length(letting_date)-3) AS INTEGER) <= {year_end}"
    
    # Price each item
    results_summary = {
        'items_requested': len(items_to_price),
        'items_priced': 0,
        'items_not_found': 0,
        'total_value': 0
    }
    
    for item in items_to_price:
        # Query for weighted average price - WINNING BIDS ONLY
        query = f"""
            SELECT 
                item_number,
                item_description,
                unit,
                SUM(extension) / NULLIF(SUM(quantity), 0) as weighted_avg_price,
                COUNT(*) as bid_count,
                MIN(unit_price) as min_price,
                MAX(unit_price) as max_price
            FROM bids
            WHERE item_number = ?
            AND unit_price > 0
            AND quantity > 0
            AND is_winner = 'Y'
            {district_clause}
            {year_clause}
            GROUP BY item_number
        """
        
        params = [item['item_number']]
        if district_list:
            params.extend(district_list)
        
        cursor.execute(query, params)
        result = cursor.fetchone()
        
        row_num = item['row']
        
        if result and result['weighted_avg_price']:
            # Fill in the data
            price = result['weighted_avg_price']
            extension = price * item['quantity'] if item['quantity'] else 0
            
            # Column B: Description (if empty)
            if not ws.cell(row=row_num, column=2).value:
                ws.cell(row=row_num, column=2).value = result['item_description']
            
            # Column D: Unit (if empty)
            if not ws.cell(row=row_num, column=4).value:
                ws.cell(row=row_num, column=4).value = result['unit']
            
            # Column E: Unit Price
            ws.cell(row=row_num, column=5).value = round(price, 2)
            ws.cell(row=row_num, column=5).number_format = '$#,##0.00'
            
            # Column F: Extension
            ws.cell(row=row_num, column=6).value = round(extension, 2)
            ws.cell(row=row_num, column=6).number_format = '$#,##0.00'
            
            # Column G: Bid count (for reference)
            ws.cell(row=row_num, column=7).value = result['bid_count']
            
            results_summary['items_priced'] += 1
            results_summary['total_value'] += extension
        else:
            # Item not found - mark it
            ws.cell(row=row_num, column=5).value = "NOT FOUND"
            ws.cell(row=row_num, column=5).font = Font(color="FF0000", italic=True)
            results_summary['items_not_found'] += 1
    
    conn.close()
    
    # Add/update headers if they exist
    if header_row > 0:
        headers = ['Item Number', 'Description', 'Quantity', 'Unit', 'Unit Price', 'Extension', 'Bid Count']
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
    
    # Add summary at bottom
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1).value = "PRICING SUMMARY"
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    ws.cell(row=summary_row + 1, column=1).value = "Items Requested:"
    ws.cell(row=summary_row + 1, column=2).value = results_summary['items_requested']
    
    ws.cell(row=summary_row + 2, column=1).value = "Items Priced:"
    ws.cell(row=summary_row + 2, column=2).value = results_summary['items_priced']
    
    ws.cell(row=summary_row + 3, column=1).value = "Items Not Found:"
    ws.cell(row=summary_row + 3, column=2).value = results_summary['items_not_found']
    
    ws.cell(row=summary_row + 4, column=1).value = "Total Estimated Value:"
    ws.cell(row=summary_row + 4, column=2).value = results_summary['total_value']
    ws.cell(row=summary_row + 4, column=2).number_format = '$#,##0.00'
    
    if district_list:
        ws.cell(row=summary_row + 5, column=1).value = "Districts Used:"
        ws.cell(row=summary_row + 5, column=2).value = ', '.join(district_list)
    
    if year_start or year_end:
        ws.cell(row=summary_row + 6, column=1).value = "Year Range:"
        ws.cell(row=summary_row + 6, column=2).value = f"{year_start or 'All'} - {year_end or 'All'}"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as downloadable file
    filename = f"priced_estimate_{file.filename}" if file.filename else "priced_estimate.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "X-Items-Priced": str(results_summary['items_priced']),
            "X-Items-Not-Found": str(results_summary['items_not_found']),
            "X-Total-Value": str(round(results_summary['total_value'], 2))
        }
    )


@router.get("/estimator/template")
@limiter.limit("10/minute")
async def get_estimator_template(request: Request):
    """
    Download a blank template Excel file for the estimator tool.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estimate Items"
    
    # Headers
    headers = ['Item Number', 'Description', 'Quantity', 'Unit', 'Unit Price', 'Extension', 'Bid Count']
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Example rows
    examples = [
        ('40603080', '', 1000, 'SQ YD', '', '', ''),
        ('60105100', '', 500, 'TON', '', '', ''),
        ('78004000', '', 200, 'LIN FT', '', '', ''),
    ]
    
    for row_num, example in enumerate(examples, 2):
        for col_num, value in enumerate(example, 1):
            ws.cell(row=row_num, column=col_num).value = value if value else None
    
    # Instructions
    ws.cell(row=7, column=1).value = "Instructions:"
    ws.cell(row=7, column=1).font = Font(bold=True)
    ws.cell(row=8, column=1).value = "1. Enter IDOT item numbers in Column A"
    ws.cell(row=9, column=1).value = "2. Enter quantities in Column C"
    ws.cell(row=10, column=1).value = "3. Upload this file to get weighted average prices"
    ws.cell(row=11, column=1).value = "4. Maximum 300 items per upload"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=estimator_template.xlsx"}
    )
